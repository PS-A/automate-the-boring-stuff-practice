# Excel-to-CSV Converter

import os, openpyxl, csv
from pathlib import Path
from typing import cast
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.chartsheet.chartsheet import Chartsheet

for excel_file in os.listdir(os.getcwd()):
    if not excel_file.endswith(".xlsx"):
        continue  # skip non-xlsx files
    print(f"Converting {excel_file} from .xlsx to .csv...")
    wb = openpyxl.load_workbook(excel_file, data_only=True)

    # Loop through every sheet in the workbook.
    for sheet_name in wb.sheetnames:
        ws_any = wb[sheet_name]
        if isinstance(ws_any, Chartsheet):
            print(f"  Skipping chartsheet: {sheet_name}")
            continue

        sheet = cast(Worksheet, ws_any)

        out_path = Path(excel_file).with_suffix("")
        csv_path = f"{out_path.name}_{sheet_name}.csv"

        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in sheet.iter_rows(values_only=True):
                writer.writerow(row)