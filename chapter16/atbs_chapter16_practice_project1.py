# Excel-to-CSV Converter

import os, openpyxl, csv
from pathlib import Path
from openpyxl.utils.cell import range_boundaries

for excelFile in os.listdir(os.getcwd()):
    if not excelFile.endswith('.xlsx'):
        continue    # skip non-xlsx files
    print(f"Converting {excelFile} from .xlsx to .csv...")
    wb = openpyxl.load_workbook(excelFile)

    # Loop through every sheet in the workbook.
    for sheetName in wb.sheetnames:
        sheet = wb[sheetName]
        min_col, min_row, max_col, max_row = range_boundaries(sheet.calculate_dimension())
        # Create the CSV
        outputFile = open(Path(excelFile).stem + f"_{sheetName}.csv", 'w', newline='')
        outputWriter = csv.writer(outputFile)
        # Loop through every row in the sheet.
        for rowNum in range(min_row, max_row + 1): #type:ignore
            rowData = []
            for colNum in range(min_col, max_col + 1): #type:ignore
                rowData.append(sheet.cell(row=rowNum, column=colNum).value)
            outputWriter.writerow(rowData)
        outputFile.close()