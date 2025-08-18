# Spreadsheet Cell Inverter

import openpyxl


def invert_cells(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    excel_data = []

    for row in sheet.iter_rows(values_only=True):  # type:ignore
        excel_data.append(list(row))
    transposed_data = list(zip(*excel_data))

    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active
    for i, row in enumerate(transposed_data, start=1):
        for j, value in enumerate(row, start=1):
            new_sheet.cell(row=i, column=j).value = value  # type:ignore
    new_wb.save(f"new_{filename}")


invert_cells("Invert Cells Test.xlsx")
