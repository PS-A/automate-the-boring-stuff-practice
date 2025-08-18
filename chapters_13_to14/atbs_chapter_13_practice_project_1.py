# Multiplication Table Maker
#! python3

import openpyxl
from openpyxl.styles import Font


def multiplication_table(n):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Table"  # type:ignore

    bold_font = Font(bold=True)

    # Fill headers (row 1 and column A)
    for i in range(1, n + 1):
        sheet.cell(row=1, column=i + 1, value=i).font = bold_font  # type:ignore
        sheet.cell(row=i + 1, column=1, value=i).font = bold_font  # type:ignore

    # Fill multiplication table
    for row in range(1, n + 1):
        for col in range(1, n + 1):
            sheet.cell(row=row + 1, column=col + 1, value=row * col)  # type:ignore

    wb.save("Multiplication Table.xlsx")


multiplication_table(5)
