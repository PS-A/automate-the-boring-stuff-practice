# Read Census Excel
"""
#! python3
# Tabulates population and number of census tracts for each county.

import openpyxl, pprint
print('Opening workbook...')
wb = openpyxl.load_workbook('censuspopdata.xlsx')
sheet = wb['Population by Census Tract']
countyData = {}

# TODO: Fill in countyData with each county's population and tracts.
print('Reading rows...')
for row in range(2, sheet.max_row + 1):
    # Each row in the spreadsheet has data for one census tract.
    state  = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop    = sheet['D' + str(row)].value

    # Make sure the key for this state exists.
    countyData.setdefault(state, {})
    # Make sure the key for this county in this state exists.
    countyData[state].setdefault(county, {'tracts': 0, 'pop': 0})

    # Each row represents one census tract, so increment by one.
    countyData[state][county]['tracts'] += 1
    # Increase the county pop by the pop in this census tract.
    countyData[state][county]['pop'] += int(pop)

# Open a new text file and write the contents of countyData to it.
print('Writing results...')
resultFile = open('census2010.py', 'w')
resultFile.write('allData = ' + pprint.pformat(countyData))
resultFile.close()
print('Done.')
"""

# Update Produce
""""""
#! python3
# Corrects costs in produce sales spreadsheet.

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

# The produce types and their updated prices
PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}

# Loop through the rows and update the prices.
for rowNum in range(2, sheet.max_row):    # skip the first row
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
            sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName] # type: ignore

wb.save('updatedProduceSales.xlsx')